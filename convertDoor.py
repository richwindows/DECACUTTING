import pandas as pd
from openpyxl import load_workbook
import csv
import os
import shutil

def get_material_color_suffix(note_value):
    """
    根据note字段内容确定材料颜色后缀
    默认返回-WH，如果note中包含Black返回-BL，包含Almond返回-AL
    """
    if note_value:
        note_str = str(note_value).upper()
        if 'BLACK' in note_str:
            return '-BL'
        elif 'ALMOND' in note_str:
            return '-AL'
    return '-WH'

def get_material_color_suffix(color_value):
    """
    根据颜色值确定材料颜色后缀
    默认返回-WH，如果颜色为Black返回-BL，为Almond返回-AL
    """
    if color_value == 'Black':
        return '-BL'
    elif color_value == 'Almond':
        return '-AL'
    else:
        return '-WH'

def process_file(xlsm_file):
    # 获取文件名（不包含路径）
    file_name = os.path.basename(xlsm_file)
    # 生成输出CSV文件路径
    csv_file = os.path.join(os.path.dirname(xlsm_file), f"{os.path.splitext(file_name)[0]}_CutFrame.csv")

    # 定义CSV文件的header
    header = [
        "Batch No", "Order No", "Order Item", "Material Name", "Cutting ID", "Pieces ID", "Length",
        "Angles", "Qty", "Bin No", "Cart No", "Position", "Label Print", "Barcode No", "PO No",
        "Style", "Frame", "Product Size", "Color", "Grid", "Glass", "Argon", "Painting",
        "Product Date", "Balance", "Shift", "Ship date", "Note", "Customer"
    ]

    # 创建一个临时CSV文件
    temp_csv_file = os.path.join(os.path.dirname(xlsm_file), f"temp_{os.path.splitext(file_name)[0]}_CutFrame.csv")

    with open(temp_csv_file, "w", encoding='utf-8', newline="") as csvfile:
        writer = csv.writer(csvfile, dialect="excel")
        writer.writerow(header)

    
         # 打开xlsm文件
        workbook = load_workbook(xlsm_file)
        frame_sheet = workbook["Frame,Sash"]
        info_sheet = workbook["Info"]

        batch_value = frame_sheet["B2"].value

        # 定义 Length_Line 变量
        Length_Line = 2

        # 创建字典来存储相同Material Name的行
        material_rows = {}

        # 从D列第4行开始遍历并写入写入130-01-
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"D{row}"].value
            door_id_value = frame_sheet[f"B{row}"].value
            door_style_value = frame_sheet[f"C{row}"].value
            #door_color_value = frame_sheet[f"O{row}"].value
            base_material_name = 'HMST130-01'
            material_pcs = '1'

            if frame_length_value is not None:

                # 在Info sheet中匹配door_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == door_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量door_customer_Value为该行第A列数据
                        door_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量door_width_value为该行第D列内容
                        door_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量door_height_value为该行第E列内容
                        door_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量door_frame_value为该行第F列内容
                        door_frame_value = info_sheet[f"F{info_row}"].value
                        # 设置变量door_glass_value为该行G列内容
                        door_glass_value = info_sheet[f"G{info_row}"].value
                        # 设置变量door_Argon_value为该行第H列内容
                        door_Argon_value = info_sheet[f"H{info_row}"].value
                        # 设置变量door_grid_value为该行第I列内容
                        door_grid_value = info_sheet[f"I{info_row}"].value
                        # 设置变量door_note_value为该行第K列内容
                        door_note_value = info_sheet[f"K{info_row}"].value
                        
                        # 根据note字段设置材料名称的颜色后缀
                        material_name = base_material_name + get_material_color_suffix(door_note_value)
                        break

                
                # 修改Material Name的值
                # if door_color_value == "Almond":
                #     material_name = str(material_name) + "-AL"
                # else:
                #     material_name = str(material_name) + "-WH"
                

                writer.writerow([""]+ [door_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [f"{frame_length_value:.6f}"] + ["V"] + [material_pcs] +[door_id_value] * 1 + [""] + ["TOP+BOT"] + [""] * 3 + [door_style_value] +[door_frame_value] + [""] * 1 + [""] + [door_grid_value] +[door_glass_value] + [door_Argon_value] + [""] *5 + [door_note_value] + [door_customer_Value])

        
        # 从H列第4行开始遍历并写入写入130-01-
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"H{row}"].value
            door_id_value = frame_sheet[f"B{row}"].value
            door_style_value = frame_sheet[f"C{row}"].value
            #door_pcs_value = frame_sheet[f"G{row}"].value
            base_material_name = 'HMST130-01'
            material_pcs = frame_sheet[f"I{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配door_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == door_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量door_customer_Value为该行第A列数据
                        door_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量door_width_value为该行第D列内容
                        door_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量door_height_value为该行第E列内容
                        door_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量door_frame_value为该行第F列内容
                        door_frame_value = info_sheet[f"F{info_row}"].value
                        # 设置变量door_glass_value为该行G列内容
                        door_glass_value = info_sheet[f"G{info_row}"].value
                        # 设置变量door_Argon_value为该行第H列内容
                        door_Argon_value = info_sheet[f"H{info_row}"].value
                        # 设置变量door_grid_value为该行第I列内容
                        door_grid_value = info_sheet[f"I{info_row}"].value
                        # 设置变量door_note_value为该行第K列内容
                        door_note_value = info_sheet[f"K{info_row}"].value
                        
                        # 根据note字段设置材料名称的颜色后缀
                        material_name = base_material_name + get_material_color_suffix(door_note_value)
                        break


                writer.writerow([""]+ [door_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [f"{frame_length_value:.6f}"] + ["V"] + [material_pcs] +[door_id_value] * 1 + [""]  + ["TOP+BOT"] + [""] * 3 + [door_style_value] +[door_frame_value] + [""] * 1 + [""] + [door_grid_value] +[door_glass_value] + [door_Argon_value] + [""] *5 + [door_note_value] + [door_customer_Value])



        # 从J列第4行开始遍历并写入写入130-01|
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"J{row}"].value
            door_id_value = frame_sheet[f"B{row}"].value
            door_style_value = frame_sheet[f"C{row}"].value
            #door_pcs_value = frame_sheet[f"G{row}"].value
            base_material_name = 'HMST130-01'
            material_pcs = frame_sheet[f"K{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配door_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == door_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量door_customer_Value为该行第A列数据
                        door_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量door_width_value为该行第D列内容
                        door_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量door_height_value为该行第E列内容
                        door_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量door_frame_value为该行第F列内容
                        door_frame_value = info_sheet[f"F{info_row}"].value
                        # 设置变量door_glass_value为该行G列内容
                        door_glass_value = info_sheet[f"G{info_row}"].value
                        # 设置变量door_Argon_value为该行第H列内容
                        door_Argon_value = info_sheet[f"H{info_row}"].value
                        # 设置变量door_grid_value为该行第I列内容
                        door_grid_value = info_sheet[f"I{info_row}"].value
                        # 设置变量door_note_value为该行第K列内容
                        door_note_value = info_sheet[f"K{info_row}"].value
                        
                        # 根据note字段设置材料名称的颜色后缀
                        material_name = base_material_name + get_material_color_suffix(door_note_value)
                        break


                writer.writerow([""]+ [door_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [f"{frame_length_value:.6f}"] + ["V"] + [material_pcs] +[door_id_value] * 1 + [""]  + ["Left+Right"] + [""] * 3 + [door_style_value] +[door_frame_value] + [""] * 1 + [""] + [door_grid_value] +[door_glass_value] + [door_Argon_value] + [""] *5 + [door_note_value] + [door_customer_Value])



        # 从E列第4行开始遍历并写入写入130-01B—
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"E{row}"].value
            door_id_value = frame_sheet[f"B{row}"].value
            door_style_value = frame_sheet[f"C{row}"].value
            #door_pcs_value = frame_sheet[f"G{row}"].value
            base_material_name = 'HMST130-01B'
            material_pcs = '1'#"frame_sheet[f"K{row}"].value"

            if frame_length_value is not None:

                # 在Info sheet中匹配door_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == door_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量door_customer_Value为该行第A列数据
                        door_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量door_width_value为该行第D列内容
                        door_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量door_height_value为该行第E列内容
                        door_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量door_frame_value为该行第F列内容
                        door_frame_value = info_sheet[f"F{info_row}"].value
                        # 设置变量door_glass_value为该行G列内容
                        door_glass_value = info_sheet[f"G{info_row}"].value
                        # 设置变量door_Argon_value为该行第H列内容
                        door_Argon_value = info_sheet[f"H{info_row}"].value
                        # 设置变量door_grid_value为该行第I列内容
                        door_grid_value = info_sheet[f"I{info_row}"].value
                        # 设置变量door_note_value为该行第K列内容
                        door_note_value = info_sheet[f"K{info_row}"].value
                        
                        # 根据note字段设置材料名称的颜色后缀
                        material_name = base_material_name + get_material_color_suffix(door_note_value)

                            
                            # 如果door_frame_value的值是'Retrofit-4'，则将material_pcs设为'2'
                        if door_frame_value == 'Retrofit-4':
                            material_pcs = '2'

                        
                        break


                            


                writer.writerow([""]+ [door_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [f"{frame_length_value:.6f}"] + ["V"] + [material_pcs] +[door_id_value] * 1 + [""] + ["TOP+BOT"] + [""] * 3 + [door_style_value] +[door_frame_value] + [""] * 1 + [""] + [door_grid_value] +[door_glass_value] + [door_Argon_value] + [""] *5 + [door_note_value] + [door_customer_Value])



        # 从F列第4行开始遍历并写入写入130-01B|
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"F{row}"].value
            door_id_value = frame_sheet[f"B{row}"].value
            door_style_value = frame_sheet[f"C{row}"].value
            #door_pcs_value = frame_sheet[f"G{row}"].value
            base_material_name = 'HMST130-01B'
            material_pcs = frame_sheet[f"G{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配door_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == door_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量door_customer_Value为该行第A列数据
                        door_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量door_width_value为该行第D列内容
                        door_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量door_height_value为该行第E列内容
                        door_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量door_frame_value为该行第F列内容
                        door_frame_value = info_sheet[f"F{info_row}"].value
                        # 设置变量door_glass_value为该行G列内容
                        door_glass_value = info_sheet[f"G{info_row}"].value
                        # 设置变量door_Argon_value为该行第H列内容
                        door_Argon_value = info_sheet[f"H{info_row}"].value
                        # 设置变量door_grid_value为该行第I列内容
                        door_grid_value = info_sheet[f"I{info_row}"].value
                        # 设置变量door_note_value为该行第K列内容
                        door_note_value = info_sheet[f"K{info_row}"].value
                        
                        # 根据note字段设置材料名称的颜色后缀
                        material_name = base_material_name + get_material_color_suffix(door_note_value)
                        break


                writer.writerow([""]+ [door_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [f"{frame_length_value:.6f}"] + ["V"] + [material_pcs] + [door_id_value] * 1 + [""] + ["Left+Right"] + [""] * 3 + [door_style_value] +[door_frame_value] + [""] * 1 + [""] + [door_grid_value] +[door_glass_value] + [door_Argon_value] + [""] *5 + [door_note_value] + [door_customer_Value])


        # 从L列第4行开始遍历并写入写入130-02-
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"L{row}"].value
            door_id_value = frame_sheet[f"B{row}"].value
            door_style_value = frame_sheet[f"C{row}"].value
            #door_pcs_value = frame_sheet[f"G{row}"].value
            material_name = 'HMST130-02'
            material_pcs = frame_sheet[f"M{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配door_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == door_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量door_customer_Value为该行第A列数据
                        door_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量door_width_value为该行第D列内容
                        door_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量door_height_value为该行第E列内容
                        door_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量door_frame_value为该行第F列内容
                        door_frame_value = info_sheet[f"F{info_row}"].value
                        # 设置变量door_glass_value为该行G列内容
                        door_glass_value = info_sheet[f"G{info_row}"].value
                        # 设置变量door_Argon_value为该行第H列内容
                        door_Argon_value = info_sheet[f"H{info_row}"].value
                        # 设置变量door_grid_value为该行第I列内容
                        door_grid_value = info_sheet[f"I{info_row}"].value
                        # 设置变量door_note_value为该行第K列内容
                        door_note_value = info_sheet[f"K{info_row}"].value
                        break


                if material_pcs == 4:
                    writer.writerow([batch_value] + [door_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [f"{frame_length_value:.6f}"] + ["V"] + [2] + [door_id_value] + [""] * 1 + ["TOP+BOT"] + [""] * 3 + [door_style_value] + [door_frame_value] + [""] * 1 + [""] + [door_grid_value] +[door_glass_value] + [door_Argon_value] + [""] * 5 + [door_note_value] + [door_customer_Value])  
                else:
                    writer.writerow([batch_value] + [door_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [f"{frame_length_value:.6f}"] + ["V"] + [material_pcs] + [door_id_value] + [""] * 1 + ["TOP+BOT"] + [""] * 3 + [door_style_value] + [door_frame_value] + [""] * 1 + [""] + [door_grid_value] +[door_glass_value] + [door_Argon_value] + [""] * 5 + [door_note_value] + [door_customer_Value])
        


# 在临时CSV文件写入完成后，添加排序逻辑
    # 关闭workbook以释放文件句柄
    workbook.close()
    
    df = pd.read_csv(temp_csv_file)
    df_sorted = df.sort_values(by=['Material Name', 'Qty', 'Length'], ascending=[True, True, False])
    
    # 将排序后的数据写入最终的CSV文件
    # df_sorted.to_csv(csv_file, index=False, encoding='utf-8-sig')
    
    # 删除临时文件
    os.remove(temp_csv_file)
    
    print("转换和排序完成！")
    
    print(df_sorted)
    
    # 复制处理后的数据到新文件
    # copy_file = os.path.join(os.path.dirname(csv_file), f"Copy_{os.path.basename(csv_file)}")
    # shutil.copy2(csv_file, copy_file)
    
    # print(f"数据已复制到 {copy_file}")

    print(df_sorted)

    return df_sorted,csv_file  # 返回处理后的DataFrame
