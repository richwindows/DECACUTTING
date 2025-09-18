import pandas as pd
from openpyxl import load_workbook
import csv
import os
import shutil

def get_material_color_suffix(color_value):
    """
    根据Color字段内容确定材料颜色后缀
    默认返回-WH，如果Color中包含Black返回-BL，包含Almond返回-AL
    """
    if color_value:
        color_str = str(color_value).upper()
        if 'BLACK' in color_str:
            return '-BL'
        elif 'ALMOND' in color_str:
            return '-AL'
    return '-WH'

def process_file(xlsm_file):
    # 获取文件名（不包含路径）
    file_name = os.path.basename(xlsm_file)
    # 生成输出CSV文件路径
    csv_file = os.path.join(os.path.dirname(xlsm_file), f"{os.path.splitext(file_name)[0]}_CutFrame.csv")

    # 定义CSV文件的header
    header = [
        "Batch No", "Order No", "Order Item", "Material Name", "Cutting ID", "Pieces ID", "Length",
        "Angles", "Qty", "Bin No", "Cart No", "Position", "Label Print", "Barcode No", "PO No",
        "Style", "Frame", "Product Size", "Color", "Grid", "Glass", "Argon", "Painting",
        "Product Date", "Balance", "Shift", "Ship date", "Note", "Customer"
    ]

    # 创建一个临时CSV文件
    temp_csv_file = os.path.join(os.path.dirname(xlsm_file), f"temp_{os.path.splitext(file_name)[0]}_CutFrame.csv")

    with open(temp_csv_file, "w", encoding='utf-8', newline="") as csvfile:
        writer = csv.writer(csvfile, dialect="excel")
        writer.writerow(header)

        # 打开xlsm文件
        workbook = load_workbook(xlsm_file, read_only=True, data_only=True)
        frame_sheet = workbook["Frame"]
        info_sheet = workbook["Info"]
        sash_sheet = workbook["Sash"]

        batch_value = frame_sheet["B2"].value

        # ... (保留原有的处理逻辑，使用temp_csv_file替代csv_file)

        # 创建字典来存储相同Material Name的行
        material_rows = {}

        # 从C列第4行开始遍历并写入写入8202B -
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"C{row}"].value
            window_id_value = frame_sheet[f"A{row}"].value
            window_style_value = frame_sheet[f"B{row}"].value
            window_color_value = frame_sheet[f"O{row}"].value
            base_material_name = 'HMST82-02B'
            material_pcs = frame_sheet[f"D{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)
                
                
                

                writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1+ ["TOP+BOT"] + [""] * 3 + [window_style_value] +[window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] +[window_glass_value] + [window_Argon_value] + [""] *5 + [window_note_value] + [window_customer_Value])

        
        # 从E列第4行开始遍历并写入写入8202B |
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"E{row}"].value
            window_id_value = frame_sheet[f"A{row}"].value
            window_style_value = frame_sheet[f"B{row}"].value
            window_color_value = frame_sheet[f"O{row}"].value
            base_material_name = 'HMST82-02B'
            material_pcs = frame_sheet[f"F{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1+ ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] +[window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] +[window_glass_value] + [window_Argon_value] + [""] *5 + [window_note_value] + [window_customer_Value])


        
        # 从G列第4行开始遍历并写入写入82-10 -
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"G{row}"].value
            window_id_value = frame_sheet[f"A{row}"].value
            window_style_value = frame_sheet[f"B{row}"].value
            window_color_value = frame_sheet[f"O{row}"].value
            base_material_name = 'HMST82-10'
            material_pcs = frame_sheet[f"H{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1+ ["TOP+BOT"] + [""] * 3 + [window_style_value] +[window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] +[window_glass_value] + [window_Argon_value] + [""] *5 + [window_note_value] + [window_customer_Value])


        # 从I列第4行开始遍历并写入写入8210 |
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"I{row}"].value
            window_id_value = frame_sheet[f"A{row}"].value
            window_style_value = frame_sheet[f"B{row}"].value
            window_color_value = frame_sheet[f"O{row}"].value
            base_material_name = 'HMST82-10'
            material_pcs = frame_sheet[f"J{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] +[window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] +[window_glass_value] + [window_Argon_value] + [""] *5 + [window_note_value] + [window_customer_Value])


        # 从K列第4行开始遍历并写入写入82-01 -
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"K{row}"].value
            window_id_value = frame_sheet[f"A{row}"].value
            window_style_value = frame_sheet[f"B{row}"].value
            window_color_value = frame_sheet[f"O{row}"].value
            base_material_name = 'HMST82-01'
            material_pcs = frame_sheet[f"L{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1+ ["TOP+BOT"] + [""] * 3 + [window_style_value] +[window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] +[window_glass_value] + [window_Argon_value] + [""] *5 + [window_note_value] + [window_customer_Value])


        # 从M列第4行开始遍历并写入写入8201 |
        for row in range(4, frame_sheet.max_row + 1):
            frame_length_value = frame_sheet[f"M{row}"].value
            window_id_value = frame_sheet[f"A{row}"].value
            window_style_value = frame_sheet[f"B{row}"].value
            window_color_value = frame_sheet[f"O{row}"].value
            base_material_name = 'HMST82-01'
            material_pcs = frame_sheet[f"N{row}"].value

            if frame_length_value is not None:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] +[window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] +[window_glass_value] + [window_Argon_value] + [""] *5 + [window_note_value] + [window_customer_Value])


        
        #sash

         # 从C列第4行开始遍历并写入82-03 一数据
        for row in range(4, sash_sheet.max_row + 1):
            frame_length_value = sash_sheet[f"C{row}"].value
            window_id_value = sash_sheet[f"A{row}"].value
            window_style_value = sash_sheet[f"B{row}"].value
            window_color_value = sash_sheet[f"H{row}"].value
            base_material_name = 'HMST82-03'
            material_pcs = sash_sheet[f"D{row}"].value

            if frame_length_value is not None and float(frame_length_value) >14:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                if material_pcs == 4:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["TOP+BOT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["TOP+BOT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])  
                else:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1 + ["TOP+BOT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
        
         # 从C列第4行开始遍历并写入82-03 |数据
        for row in range(4, sash_sheet.max_row + 1):
            frame_length_value = sash_sheet[f"E{row}"].value
            window_id_value = sash_sheet[f"A{row}"].value
            window_style_value = sash_sheet[f"B{row}"].value
            window_color_value = sash_sheet[f"H{row}"].value
            base_material_name = 'HMST82-03'
            material_pcs = sash_sheet[f"F{row}"].value

            if frame_length_value is not None and float(frame_length_value) >14:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                if material_pcs == 4:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])  
                else:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
        

        # 从C列第4行开始遍历并写入82-05 |数据
        for row in range(4, sash_sheet.max_row + 1):
            frame_length_value = sash_sheet[f"G{row}"].value
            window_id_value = sash_sheet[f"A{row}"].value
            window_style_value = sash_sheet[f"B{row}"].value
            window_color_value = sash_sheet[f"H{row}"].value    
            base_material_name = 'HMST82-05'
            material_pcs = sash_sheet[f"H{row}"].value

            if frame_length_value is not None and float(frame_length_value) >14:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                if material_pcs == 4:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])  
                else:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
        


        # 从C列第4行开始遍历并写入82-04 一 据
        for row in range(4, sash_sheet.max_row + 1):
            frame_length_value = sash_sheet[f"I{row}"].value
            window_id_value = sash_sheet[f"A{row}"].value
            window_style_value = sash_sheet[f"B{row}"].value
            window_color_value = sash_sheet[f"H{row}"].value
            base_material_name = 'HMST82-04'
            material_pcs = sash_sheet[f"J{row}"].value

            if frame_length_value is not None and float(frame_length_value) >14:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                if material_pcs == 4:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["TOP+BOT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["TOP+BOT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])  
                else:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1 + ["TOP+BOT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
        


        # 从C列第4行开始遍历并写入82-04 |数据
        for row in range(4, sash_sheet.max_row + 1):
            frame_length_value = sash_sheet[f"K{row}"].value
            window_id_value = sash_sheet[f"A{row}"].value
            window_style_value = sash_sheet[f"B{row}"].value
            window_color_value = sash_sheet[f"O{row}"].value
            base_material_name = 'HMST82-04'
            material_pcs = sash_sheet[f"L{row}"].value

            if frame_length_value is not None and float(frame_length_value) >14:

                # 在Info sheet中匹配window_id_value，并获取相应的行数
                for info_row in range(2, info_sheet.max_row + 1):
                    info_id = info_sheet[f"B{info_row}"].value
                    if info_id == window_id_value:
                        # 设置变量Info_ID_Line为匹配行数
                        Info_ID_Line = info_row
                        # 设置变量window_customer_Value为该行第A列数据
                        window_customer_Value = info_sheet[f"A{info_row}"].value
                        # 设置变量window_width_value为该行第D列内容
                        window_width_value = info_sheet[f"D{info_row}"].value
                        # 设置变量window_height_value为该行第E列内容
                        window_height_value = info_sheet[f"E{info_row}"].value
                        # 设置变量window_frame_value为该行第G列内容
                        window_frame_value = info_sheet[f"G{info_row}"].value
                        # 设置变量window_glass_value为该行H列内容
                        window_glass_value = info_sheet[f"H{info_row}"].value
                        # 设置变量window_Argon_value为该行第I列内容
                        window_Argon_value = info_sheet[f"I{info_row}"].value
                        # 设置变量window_grid_value为该行第J列内容
                        window_grid_value = info_sheet[f"J{info_row}"].value
                        # 设置变量window_color_value为该行第K列内容
                        window_color_value = info_sheet[f"K{info_row}"].value
                        # 设置变量window_note_value为该行第L列内容
                        window_note_value = info_sheet[f"L{info_row}"].value
                        break

                # 根据Color字段设置材料名称的颜色后缀
                material_name = base_material_name + get_material_color_suffix(window_color_value)

                if material_pcs == 4:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [2] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])  
                else:
                    writer.writerow([batch_value] + [window_id_value] * 1 + ["1"] + [material_name] + [""] * 2 + [frame_length_value] + ["V"] + [material_pcs] + [window_id_value] + [""] * 1 + ["LEFT+RIGHT"] + [""] * 3 + [window_style_value] + [window_frame_value] + [""] * 1 + [window_color_value] + [window_grid_value] + [window_glass_value] + [window_Argon_value] + [""] * 5 + [window_note_value] + [window_customer_Value])
        




# 在临时CSV文件写入完成后，添加排序逻辑
    df = pd.read_csv(temp_csv_file)
    df_sorted = df.sort_values(by=['Material Name', 'Qty', 'Length'], ascending=[True, True, False])
    
    # 将排序后的数据写入最终的CSV文件
    # df_sorted.to_csv(csv_file, index=False, encoding='utf-8-sig')
    
    # 删除临时文件
    os.remove(temp_csv_file)
    
    print("转换和排序完成！")
    
    print(df_sorted)
    
    # 复制处理后的数据到新文件
    # copy_file = os.path.join(os.path.dirname(csv_file), f"Copy_{os.path.basename(csv_file)}")
    # shutil.copy2(csv_file, copy_file)
    
    # print(f"数据已复制到 {copy_file}")

    return df_sorted,csv_file  # 返回处理后的DataFrame